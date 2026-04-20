#!/usr/bin/env python
"""table.py — CLI for the Table sheet of LMNP.xlsx

Commands:
  list            List entries (filterable; --format table|tsv)
  delete-range    Delete rows in a date range (--dry-run supported)
  append-csv      Insert rows from a CSV in date order (--dry-run supported)
  summary         Totals by Nature for a year
  justificatif-max  Show current max justificatif number

CSV schema for append-csv:
  date,moyen,compte,nature,commentaire,debit,credit,justificatif
  (date = YYYY-MM-DD; debit/credit = positive float or empty;
   justificatif quoted if it contains a comma, e.g. "144,45")
"""
import argparse
import csv
import sys
from datetime import datetime

# Windows cmd/PowerShell default to cp1252; force UTF-8 so accented chars print cleanly.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

WORKBOOK_DEFAULT = "LMNP.xlsx"
SHEET = "Table"
DATE_FMT = "mm-dd-yy"

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def load(path):
    wb = openpyxl.load_workbook(path)
    return wb, wb[SHEET]


def iter_data_rows(ws):
    """Yield (row_idx, date_or_None, row_values[0:8]) for non-empty rows."""
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=8, values_only=True), 2):
        if any(v is not None for v in row):
            yield i, row[0], row


def parse_date(s):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    raise argparse.ArgumentTypeError(f"Unrecognised date: {s!r}  (expected YYYY-MM-DD)")


def _fmt_val(v, width=None):
    s = "" if v is None else str(v)
    return s if width is None else s[:width].ljust(width)


def write_row(ws, idx, data):
    """Write one data tuple to a worksheet row (columns A–H)."""
    d, moyen, compte, nature, comment, debit, credit, just = data
    ws.cell(idx, 1).value = d
    ws.cell(idx, 1).number_format = DATE_FMT
    ws.cell(idx, 2).value = moyen or None
    ws.cell(idx, 3).value = compte or None
    ws.cell(idx, 4).value = nature or None
    ws.cell(idx, 5).value = comment or None
    ws.cell(idx, 6).value = float(debit) if debit else None
    ws.cell(idx, 7).value = float(credit) if credit else None
    ws.cell(idx, 8).value = just or None


# ---------------------------------------------------------------------------
# subcommands
# ---------------------------------------------------------------------------

def cmd_list(args, wb, ws):
    rows = []
    for idx, d, row in iter_data_rows(ws):
        if args.year and (not d or d.year != args.year):
            continue
        if args.nature and row[3] != args.nature:
            continue
        if args.from_date and (not d or d < args.from_date):
            continue
        if args.to_date and (not d or d > args.to_date):
            continue
        rows.append((idx, d, row))

    if args.format == "tsv":
        print("\t".join(["row", "date", "moyen", "compte", "nature",
                          "commentaire", "debit", "credit", "justificatif"]))
        for idx, d, row in rows:
            ds = d.strftime("%Y-%m-%d") if d else ""
            parts = [str(idx), ds] + [str(v) if v is not None else "" for v in row[1:8]]
            print("\t".join(parts))
    else:
        # human-readable fixed-width table
        print(f"{'#':>4}  {'date':10}  {'moyen':12}  {'nature':13}  "
              f"{'commentaire':36}  {'debit':>9}  {'credit':>9}  justificatif")
        print("-" * 115)
        for idx, d, row in rows:
            ds = d.strftime("%Y-%m-%d") if d else "?"
            debit  = f"{row[5]:.2f}" if row[5] is not None else ""
            credit = f"{row[6]:.2f}" if row[6] is not None else ""
            print(f"{idx:4d}  {ds:10}  {_fmt_val(row[1], 12)}  {_fmt_val(row[3], 13)}  "
                  f"{_fmt_val(row[4], 36)}  {debit:>9}  {credit:>9}  {row[7] or ''}")


def cmd_delete_range(args, wb, ws):
    to_delete = [
        idx for idx, d, _ in iter_data_rows(ws)
        if d and args.from_date <= d <= args.to_date
    ]
    label = "[DRY RUN] " if args.dry_run else ""
    print(f"{label}Found {len(to_delete)} rows "
          f"({args.from_date.date()} → {args.to_date.date()})")
    if args.dry_run:
        for idx in to_delete:
            row = list(ws.iter_rows(min_row=idx, max_row=idx, max_col=8, values_only=True))[0]
            d = row[0]
            ds = d.strftime("%Y-%m-%d") if d else "?"
            print(f"  row {idx:4d}  {ds}  {row[3] or ''}  {row[4] or ''}")
        return

    for idx in reversed(to_delete):
        ws.delete_rows(idx)
    wb.save(args.workbook)
    print(f"Deleted {len(to_delete)} rows. Saved {args.workbook}.")


def cmd_append_csv(args, wb, ws):
    with open(args.file, newline="", encoding="utf-8") as f:
        raw = list(csv.DictReader(f))

    parsed = sorted(
        [
            (
                parse_date(r["date"]),
                r.get("moyen", ""),
                r.get("compte", ""),
                r.get("nature", ""),
                r.get("commentaire", ""),
                r.get("debit", ""),
                r.get("credit", ""),
                r.get("justificatif", ""),
            )
            for r in raw
        ],
        key=lambda x: x[0],
    )

    if not parsed:
        print("No rows to append.")
        return

    first_date = parsed[0][0]

    # Find last existing row with date <= first_date.
    # insert_after starts at 1 (= after header, before all data).
    insert_after = 1
    for idx, d, _ in iter_data_rows(ws):
        if d and d <= first_date:
            insert_after = idx
        elif d and d > first_date:
            break

    n = len(parsed)
    label = "[DRY RUN] " if args.dry_run else ""
    print(f"{label}Inserting {n} rows after row {insert_after} "
          f"(last date <= {first_date.date()})")

    if args.dry_run:
        for data in parsed:
            d = data[0]
            print(f"  {d.strftime('%Y-%m-%d')}  {data[3]:13}  {data[4][:35]}  "
                  f"debit={data[5] or '—':>8}  credit={data[6] or '—':>8}  just={data[7]}")
        return

    ws.insert_rows(insert_after + 1, n)
    for j, data in enumerate(parsed):
        write_row(ws, insert_after + 1 + j, data)
    wb.save(args.workbook)
    print(f"Saved {args.workbook}.")


def cmd_summary(args, wb, ws):
    totals = {}
    for _, d, row in iter_data_rows(ws):
        if args.year and (not d or d.year != args.year):
            continue
        nat = row[3] or "(none)"
        t = totals.setdefault(nat, [0, 0.0, 0.0])
        t[0] += 1
        t[1] += row[5] or 0
        t[2] += row[6] or 0

    year_label = str(args.year) if args.year else "all years"
    print(f"\nSummary — {year_label}")
    print(f"{'Nature':15}  {'N':>5}  {'Débit':>12}  {'Crédit':>12}")
    print("-" * 50)
    for nat, (n, d, c) in sorted(totals.items()):
        print(f"{nat:15}  {n:>5d}  {d:>12.2f}  {c:>12.2f}")
    grand_d = sum(v[1] for v in totals.values())
    grand_c = sum(v[2] for v in totals.values())
    print("-" * 50)
    print(f"{'TOTAL':15}  {'':>5}  {grand_d:>12.2f}  {grand_c:>12.2f}")


def cmd_jmax(args, wb, ws):
    mx = 0
    for _, _, row in iter_data_rows(ws):
        j = row[7]
        if j:
            for part in str(j).split(","):
                try:
                    mx = max(mx, int(part.strip()))
                except ValueError:
                    pass
    print(f"Max justificatif: {mx}  ->  next: {mx + 1}")


# ---------------------------------------------------------------------------
# arg parsing
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(
        prog="table.py",
        description="Read/write the Table sheet of LMNP.xlsx",
    )
    p.add_argument("--workbook", default=WORKBOOK_DEFAULT,
                   help=f"Path to workbook (default: {WORKBOOK_DEFAULT})")
    sub = p.add_subparsers(dest="cmd", required=True)

    # list
    sl = sub.add_parser("list", help="List Table entries")
    sl.add_argument("--year", type=int)
    sl.add_argument("--nature")
    sl.add_argument("--from", dest="from_date", type=parse_date, metavar="YYYY-MM-DD")
    sl.add_argument("--to",   dest="to_date",   type=parse_date, metavar="YYYY-MM-DD")
    sl.add_argument("--format", choices=["table", "tsv"], default="table")

    # delete-range
    sd = sub.add_parser("delete-range", help="Delete rows in a date range")
    sd.add_argument("--from", dest="from_date", required=True,
                    type=parse_date, metavar="YYYY-MM-DD")
    sd.add_argument("--to",   dest="to_date",   required=True,
                    type=parse_date, metavar="YYYY-MM-DD")
    sd.add_argument("--dry-run", action="store_true")

    # append-csv
    sa = sub.add_parser("append-csv",
                        help="Insert rows from a CSV in date order")
    sa.add_argument("file", help="CSV file (header: date,moyen,compte,nature,"
                                 "commentaire,debit,credit,justificatif)")
    sa.add_argument("--dry-run", action="store_true")

    # summary
    ss = sub.add_parser("summary", help="Totals by Nature")
    ss.add_argument("--year", type=int)

    # justificatif-max
    sub.add_parser("justificatif-max",
                   help="Show current max justificatif number")

    args = p.parse_args()
    wb, ws = load(args.workbook)

    dispatch = {
        "list":             cmd_list,
        "delete-range":     cmd_delete_range,
        "append-csv":       cmd_append_csv,
        "summary":          cmd_summary,
        "justificatif-max": cmd_jmax,
    }
    dispatch[args.cmd](args, wb, ws)


if __name__ == "__main__":
    main()
